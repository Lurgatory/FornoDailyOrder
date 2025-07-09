import pandas as pd
import os
from openpyxl import load_workbook
from config import (
    dd_gl_ofz_file_path,
    dd_daily_order_folder_path,
    dd_result_folder_path,
    dd_pending_invoice_file_path,
    am_csv_template_path,
)
from utils.file_utils import combine_files
from utils.dd_csv_export import export_gl_csv, export_ofz_csv, export_am_csv

all_daily_orders_df = pd.DataFrame()
all_loc_df = pd.DataFrame()
all_pending_invoice_df = pd.DataFrame()
all_big_dealers_df = pd.DataFrame()
all_gl_df = pd.DataFrame()
all_ofz_df = pd.DataFrame()
am_csv_template_df = pd.read_csv(am_csv_template_path, nrows=0)
am_csv_template_column = am_csv_template_df.columns.tolist()
all_am_df = pd.DataFrame(columns=am_csv_template_column)

os.makedirs(dd_result_folder_path, exist_ok=True)

csv_files = [
    f
    for f in os.listdir(dd_daily_order_folder_path)
    if f.endswith(".csv")
    and os.path.isfile(os.path.join(dd_daily_order_folder_path, f))
]

loc_files = [
    f
    for f in os.listdir(dd_gl_ofz_file_path)
    if f.endswith(".csv") and os.path.isfile(os.path.join(dd_gl_ofz_file_path, f))
]

all_daily_orders_df = combine_files(
    csv_files, dd_daily_order_folder_path, all_daily_orders_df
)
all_loc_df = combine_files(loc_files, dd_gl_ofz_file_path, all_loc_df)

# Save OFZ and GL Dataframe
ofz_gl_df = all_loc_df[
    ~all_loc_df["note"].isna() & (all_loc_df["note"].str.strip() != "")
]

all_daily_orders_df["consigneezip"] = (
    all_daily_orders_df["consigneezip"].astype(str).str.split("-").str[0].str.zfill(5)
)
all_daily_orders_df = all_daily_orders_df[
    all_daily_orders_df["item"] != "Freight Surcharge"
]

digital_dealers = all_daily_orders_df["digital dealer"].dropna().unique()


for digital_dealer in digital_dealers:
    digital_dealer_df = all_daily_orders_df[
        all_daily_orders_df["digital dealer"] == digital_dealer
    ]
    if digital_dealer in [
        "Premium Home Source",
        "Home Outlet Direct",
        "Home Best Price dba Amazon",
    ]:
        all_big_dealers_df = pd.concat(
            [all_big_dealers_df, digital_dealer_df], ignore_index=True
        )
    else:
        all_pending_invoice_df = pd.concat(
            [all_pending_invoice_df, digital_dealer_df], ignore_index=True
        )

# Add pending invoice orders
book = load_workbook(dd_pending_invoice_file_path)
sheet_name = "PendingInvoiceOrders"

if sheet_name not in book.sheetnames:
    raise ValueError(
        f"Sheet '{sheet_name}' not found in {dd_pending_invoice_file_path}"
    )

startrow = book[sheet_name].max_row

with pd.ExcelWriter(
    dd_pending_invoice_file_path, engine="openpyxl", mode="a", if_sheet_exists="overlay"
) as writer:
    all_pending_invoice_df.to_excel(
        writer, sheet_name=sheet_name, startrow=startrow, index=False, header=False
    )

# Filter out all backorders
backorders = all_big_dealers_df.loc[
    all_big_dealers_df["pickupcity"].isna(), "memo (main)"
].unique()
non_backorders_df = all_big_dealers_df[
    ~all_big_dealers_df["memo (main)"].isin(backorders)
]

# GL and OFZ
notes = ofz_gl_df["note"].dropna().unique()

for note in notes:
    memo_list = (
        all_loc_df.loc[all_loc_df["note"] == note, "memo (main)"].dropna().unique()
    )

    matched_orders = pd.concat(
        [non_backorders_df[non_backorders_df["memo (main)"] == m] for m in memo_list],
        ignore_index=True,
    )

    if note == "OFZ":
        all_ofz_df = pd.concat([all_ofz_df, matched_orders], ignore_index=True)
    else:
        all_gl_df = pd.concat([all_gl_df, matched_orders], ignore_index=True)

    non_backorders_df = non_backorders_df[
        ~non_backorders_df["memo (main)"].isin(memo_list)
    ]

# Do filter
all_am_df = all_am_df.reindex(index=non_backorders_df.index)

export_am_csv(all_am_df, non_backorders_df)
# Do OFZ
export_ofz_csv(all_ofz_df)
# Do GL
export_gl_csv(all_gl_df)
